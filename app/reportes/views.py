# -*- coding: utf-8 -*-
from io import BytesIO

from django.http import HttpResponse, Http404
from django.views.generic import ListView,TemplateView
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table
from cartones.models import *
from salas.models import *
import time
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models import Count
from clientes.models import *
from partidas.models import *
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
from reportlab.lib.colors import HexColor
from reportlab.graphics.barcode import code128
from reportlab.lib.units import inch

class RepotSalaView(TemplateView):
    template_name = "reportSalas.html"

class RepotpremioView(TemplateView):
    template_name = "reportePremios.html"

class CombinacionView(TemplateView):
    template_name = "reporteCombinaciones.html"

class ReportSalaListView(ListView):
    model = Impresion
    template_name = "reportSalaList.html"
    desde = time.strftime("%Y-%m-%d")
    hasta = time.strftime("%Y-%m-%d")
    @method_decorator(login_required(login_url='/'))
    def dispatch(self, *args, **kwargs):
        self.desde = time.strftime("%Y-%m-%d")
        self.hasta = time.strftime("%Y-%m-%d")
        return super(ReportSalaListView, self).dispatch(*args, **kwargs)
    def post(self,request,*args,**kwargs):
        self.desde = datetime.strptime(request.POST['desde'] + " 00:00", "%m/%d/%Y %H:%M")
        self.hasta = datetime.strptime(request.POST['hasta'] + " 00:00", "%m/%d/%Y %H:%M")
        return super(ReportSalaListView, self).get(request,*args, **kwargs)
    def get_queryset(self):
        sala = Sala.objects.filter(activa='Si')
        result = []
        for s in sala:
            print s.nombre
            r = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),).aggregate(t=Count('id'))
            g1 = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),ganador_1 = True).aggregate(t=Count('id'))
            g2 = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),ganador_2 = True).aggregate(t=Count('id'))
            g3 = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),ganador_3 = True).aggregate(t=Count('id'))
            g4 = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),ganador_4 = True).aggregate(t=Count('id'))
            g6 = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),ganador_6 = True).aggregate(t=Count('id'))
            g7 = Impresion.objects.filter(partida__fecha__range=(self.desde,self.hasta),ganador_7 = True).aggregate(t=Count('id'))
            ganadores = g1['t'] + g2['t'] + g3['t'] + g4['t'] + g5['t'] + g6['t'] + g7['t'] 
            a = {'nombre': s.nombre, 'vendidos': r['t'], 'ganadores':ganadores}
            result.append(dict(a))
            print result
        return result
    def desde(self):
        return self.desde
    def hasta(self):
        return self.hasta



def generar_pdf(request):
    print "Genero el PDF"
    desde = datetime.datetime.strptime(request.POST['desde'] + " 00:00", "%m/%d/%Y %H:%M")
    hasta = datetime.datetime.strptime(request.POST['hasta'] + " 00:00", "%m/%d/%Y %H:%M")
    if request.POST:
        response = HttpResponse(content_type='application/pdf')
        pdf_name = "Reporte-Salas.pdf"  # llamado clientes
        # la linea 26 es por si deseas descargar el pdf a tu computadora
        # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
        buff = BytesIO()
        doc = SimpleDocTemplate(buff,
                                pagesize=letter,
                                rightMargin=40,
                                leftMargin=40,
                                topMargin=60,
                                bottomMargin=18,
                                )
        clientes = []
        allclientes = []
        styles = getSampleStyleSheet()
        header = Paragraph("Reporte Por Salas", styles['Heading1'])
        clientes.append(header)
        headings = ('Sala', 'Vendidos', 'Ganadores' )

        sala = Sala.objects.filter(activa='Si')
        result = []
        for s in sala:
            r = Impresion.objects.filter(partida__fecha__range=(desde,hasta),).aggregate(t=Count('id'))
            g1 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_1 = True).aggregate(t=Count('id'))
            g2 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_2 = True).aggregate(t=Count('id'))
            g3 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_3 = True).aggregate(t=Count('id'))
            g4 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_4 = True).aggregate(t=Count('id'))
            g5 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_5 = True).aggregate(t=Count('id'))
            g6 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_6 = True).aggregate(t=Count('id'))
            g7 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_7 = True).aggregate(t=Count('id'))
            g8 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_8 = True).aggregate(t=Count('id'))
            g9 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_9 = True).aggregate(t=Count('id'))
            g10 = Impresion.objects.filter(partida__fecha__range=(desde,hasta),ganador_10 = True).aggregate(t=Count('id'))
            ganadores = g1['t'] + g2['t'] + g3['t'] + g4['t'] + g5['t'] + g6['t'] + g7['t']  + g8['t']  + g9['t']  + g10['t'] 
            # a = ['nombre': s.nombre, 'vendidos': r['t'], 'ganadores':ganadores]
            a = [s.nombre, r['t'], ganadores]
            allclientes.append(list(a))
        #allclientes = [(p.username, p.email, ) for p in User.objects.all()]

        t = Table([headings] + allclientes)
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
            ]
        ))
        clientes.append(t)
        doc.build(clientes)
        response.write(buff.getvalue())
        buff.close()
        return response
    else:
        raise Http404



def ReportePCartones(request,id):
    print "Genero el PDF"
    partida = id
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte-Partidas.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    clientes = []
    allclientes = []
    styles = getSampleStyleSheet()
    header = Paragraph("Reporte de Cartones Partida Nº: " + str(partida), styles['Heading1'])
    clientes.append(header)
    headings = ('ID Carton', 'Premio', 'Fecha/Hora','Sala' )

    cartones = Impresion.objects.filter(propietario__partida_id=partida, ganador=True)
    for c in cartones:
        premio = ""
        if c.ganador_1 is True:
            premio = "Primera Linea"
        elif c.ganador_2 is True:
            premio = "Primera y Segunda Linea"
        elif c.ganador_3 is True:
            premio = "Primera y Tercera Linea"
        elif c.ganador_4 is True:
            premio = "Segunda Linea"
        elif c.ganador_5 is True:
            premio = "Segunda y Tercera Linea"
        elif c.ganador_6 is True:
            premio = "Tercera Linea"
        elif c.ganador_7 is True:
            premio = "Bingo"
        elif c.ganador_8 is True:
            premio = "Linea al Azar"
        elif c.ganador_9 is True:
            premio = "ReBingo"
        elif c.ganador_10 is True:
            premio = "Revancha"
         # a = ['nombre': s.nombre, 'vendidos': r['t'], 'ganadores':ganadores]
        a = [c.id, premio, c.ganador_at,c.propietario.sala.nombre]
        allclientes.append(list(a))
    #allclientes = [(p.username, p.email, ) for p in User.objects.all()]
    t = Table([headings] + allclientes)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
            ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
        ]
    ))
    clientes.append(t)
    doc.build(clientes)
    response.write(buff.getvalue())
    buff.close()
    return response



def combinacionCartones(request):
    print "Genero el PDF"
    partida = Partida.objects.get(pk = request.POST['id'])
    cantidadC = Impresion.objects.filter(partida_id = request.POST['id']).count()
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte-Partidas.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=A4,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    cuerpo = []
    allclientes = []
    styles = getSampleStyleSheet()
    header = Paragraph("Combincion de Cartones Partida Nº: " + str(partida.id), styles['Heading2'])
    cuerpo.append(header)
    header = Paragraph("Cantidad de cartones generdos: " + str(cantidadC), styles['Heading3'])
    cuerpo.append(header)
    if  partida:
        for c in Impresion.objects.filter(partida_id = partida.id).order_by('numeroCarton'):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
  
    doc.build(cuerpo)
    response.write(buff.getvalue())
    buff.close()
    return response

def ReportePartida(request):
    print "Genero el PDF"
    
    partida = Partida.objects.get(pk = request.POST['id'])
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte-Partidas.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=A4,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    cuerpo = []
    allclientes = []
    styles = getSampleStyleSheet()
    header = Paragraph("Reporte de Partida Nº: " + str(partida.id), styles['Heading2'])
    cuerpo.append(header)
    header = Paragraph("Organizada por: " + str(partida.organiza), styles['Heading2'])
    cuerpo.append(header)
    header = Paragraph("Fecha: " + str(partida.fecha), styles['Heading2'])
    cuerpo.append(header)
    header = Paragraph('CARTONES JUGADOS', styles['Heading3'])
    cuerpo.append(header)

    #CARTONES JUGADOS
    CartonesJugando = []
    numeroTupla = ()
    bandera = 0
    for c in Impresion.objects.filter(partida_id = partida.id, participa = 'S').order_by('numeroCarton'):
        bandera += 1
        if bandera <= 10:
            numeroTupla += (str(c.numeroCarton).zfill(5),)
            if bandera == 10:
                bandera = 0
                CartonesJugando.append(numeroTupla)
                numeroTupla = ()

    table = Table(CartonesJugando)
    table.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
            
        ]
    ))
    cuerpo.append(table)

    #CARTONES NO VENDIDOS
    header = Paragraph('CARTONES NO VENDIDOS', styles['Heading3'])
    cuerpo.append(header)
    CartonesJugando = []
    numeroTupla = ()
    bandera = 0
    padron = Impresion.objects.filter(partida_id = partida.id, participa = 'N').order_by('numeroCarton')
    
    if padron:
        for c in padron:
            bandera += 1
            if bandera <= 10:
                numeroTupla += (str(c.numeroCarton).zfill(5),)
                if bandera == 10:
                    bandera = 0
                    CartonesJugando.append(numeroTupla)
                    numeroTupla = ()
        if bandera<10:
            CartonesJugando.append(numeroTupla)
 
        table = Table(CartonesJugando)
        table.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                
            ]
        ))
        cuerpo.append(table)
    else:
        header = Paragraph('NO HAY CARTONES NO VENDIDOS', styles['Heading4'])
        cuerpo.append(header)

    #CARTONES ANULADOS
    header = Paragraph('CARTONES ANULADOS', styles['Heading3'])
    cuerpo.append(header)
    CartonesJugando = []
    numeroTupla = ()
    bandera = 0
    padron = Impresion.objects.filter(partida_id = partida.id, participa = 'A').order_by('numeroCarton')
    if padron:
        for c in padron:
            bandera += 1
            if bandera <= 10:
                numeroTupla += (str(c.numeroCarton).zfill(5),)
                if bandera == 10:
                    bandera = 0
                    CartonesJugando.append(numeroTupla)
                    numeroTupla = ()
        if bandera<10:
            CartonesJugando.append(numeroTupla)
        table = Table(CartonesJugando)
        table.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                
            ]
        ))
        cuerpo.append(table)
    else:
        header = Paragraph('NO HAY CARTONES ANULADOS', styles['Heading4'])
        cuerpo.append(header)

    #CARTONES NO JUGADOS
    header = Paragraph('CARTONES NO IMPRESOS', styles['Heading3'])
    cuerpo.append(header)
    CartonesJugando = []
    numeroTupla = ()
    bandera = 0
    padron =  Impresion.objects.filter(partida_id = partida.id, participa = 'E').order_by('numeroCarton')
    if padron:
        for c in padron:
            bandera += 1
            if bandera <= 10:
                numeroTupla += (str(c.numeroCarton).zfill(5),)
                if bandera == 10:
                    bandera = 0
                    CartonesJugando.append(numeroTupla)
                    numeroTupla = ()
        if bandera<10:
            CartonesJugando.append(numeroTupla)
 
        table = Table(CartonesJugando)
        table.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                
            ]
        ))
        cuerpo.append(table)
    else:
        header = Paragraph('NO HAY CARTONES SIN IMPRIMIR', styles['Heading4'])
        cuerpo.append(header)

    doc.build(cuerpo)
    response.write(buff.getvalue())
    buff.close()
    return response


def ReportePremios(request):
    partida = Partida.objects.get(pk = request.POST['id'])
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte-Partidas.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=A4,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    cuerpo = []
    allclientes = []
    styles = getSampleStyleSheet()
    header = Paragraph("Reporte de Premios Para la Partida Nº: " + str(partida.id), styles['Heading2'])
    cuerpo.append(header)
    header = Paragraph("Organizada por: " + str(partida.organiza), styles['Heading2'])
    cuerpo.append(header)
    header = Paragraph("Fecha: " + str(partida.fecha), styles['Heading2'])
    cuerpo.append(header)
     #PREMIO 1
    if partida.premio_1 and partida.ganador_1:
        header = Paragraph('SORTEO DEL 1ER PREMIO ' + partida.monto_1, styles['Heading3'])
        cuerpo.append(header)

        for c in Impresion.objects.filter(partida_id = partida.id, ganador_1 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            #linea 1
            lineaCarton=[]
            if c.linea_1 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
            #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_1 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
            #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_1 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)

        #balotas 1er premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '1er Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)

    #PREMIO 2
    if partida.premio_2 and partida.ganador_2:
        header = Paragraph('SORTEO DEL 2DO PREMIO ' + partida.monto_2, styles['Heading3'])
        cuerpo.append(header)
        #linea 1
        for c in Impresion.objects.filter(partida_id = partida.id, ganador_2 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            if c.linea_2 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_2 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_2 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #balotas 2do premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '2do Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)

 
    #PREMIO 3
    if partida.premio_3 and partida.ganador_3:
        header = Paragraph('SORTEO DEL 3ER PREMIO ' + partida.monto_3, styles['Heading3'])
        cuerpo.append(header)

        #linea 1
        for c in Impresion.objects.filter(partida_id = partida.id, ganador_3 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            if c.linea_3 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_3 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_3 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    print  "aaaaaaaaaaaaaaaaaaaaaaaa"
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #balotas 2do premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '3er Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)


    #PREMIO 4
    if partida.premio_4 and partida.ganador_4:
        header = Paragraph('SORTEO DEL 4to PREMIO ' + partida.monto_4, styles['Heading3'])
        cuerpo.append(header)

        #linea 1
        for c in Impresion.objects.filter(partida_id = partida.id, ganador_4 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            if c.linea_4 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_4 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_4 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #balotas 4to premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '4to Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)


    #PREMIO 5
    if partida.premio_5 and partida.ganador_5:
        header = Paragraph('SORTEO DEL 5to PREMIO ' + partida.monto_5, styles['Heading3'])
        cuerpo.append(header)

        #linea 1
        for c in Impresion.objects.filter(partida_id = partida.id, ganador_5 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            if c.linea_5 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_5 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_5 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #balotas 5to premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '5to Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)


    #PREMIO 6
    if partida.premio_5 and partida.ganador_6:
        header = Paragraph('SORTEO DEL 6to PREMIO ' + partida.monto_6, styles['Heading3'])
        cuerpo.append(header)

        #linea 1
        for c in Impresion.objects.filter(partida_id = partida.id, ganador_6 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            if c.linea_6 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_6 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_6 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #balotas 6to premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '6to Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)


    #PREMIO 7
    if partida.premio_7 and partida.ganador_7:
        header = Paragraph('SORTEO DEL 7mo PREMIO ' + partida.monto_7, styles['Heading3'])
        cuerpo.append(header)

        #linea 1
        for c in Impresion.objects.filter(partida_id = partida.id, ganador_7 = True):
            header = Paragraph('CARTON: ' + str(c.numeroCarton).zfill(5), styles['Heading4'])
            cuerpo.append(header)
            lineaCarton=[]
            if c.linea_7 == 1:
                lineaCarton.append([Paragraph('<strong>' + str(c.carton.l1_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l1_9) + '</strong>', styles['Normal'])])
            else:
                lineaCarton.append([Paragraph(str(c.carton.l1_1), styles['Normal']),Paragraph(str(c.carton.l1_2), styles['Normal']),Paragraph(str(c.carton.l1_3), styles['Normal']),Paragraph(str(c.carton.l1_4), styles['Normal']),Paragraph(str(c.carton.l1_5), styles['Normal']),Paragraph(str(c.carton.l1_6), styles['Normal']),Paragraph(str(c.carton.l1_7), styles['Normal']),Paragraph(str(c.carton.l1_8), styles['Normal']),Paragraph(str(c.carton.l1_9), styles['Normal'])])
            table = Table(lineaCarton)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)
        #linea 2
            lineaCarton=[]
            if  partida.lineasEnJuego >= 2:
                if c.linea_7 == 2:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l2_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l2_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l2_1), styles['Normal']),Paragraph(str(c.carton.l2_2), styles['Normal']),Paragraph(str(c.carton.l2_3), styles['Normal']),Paragraph(str(c.carton.l2_4), styles['Normal']),Paragraph(str(c.carton.l2_5), styles['Normal']),Paragraph(str(c.carton.l2_6), styles['Normal']),Paragraph(str(c.carton.l2_7), styles['Normal']),Paragraph(str(c.carton.l2_8), styles['Normal']),Paragraph(str(c.carton.l2_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #linea 3
            lineaCarton=[]
            if  partida.lineasEnJuego == 3:
                if c.linea_7 == 3:
                    lineaCarton.append([Paragraph('<strong>' + str(c.carton.l3_1) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_2) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_3) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_4) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_5) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_6) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_7) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_8) + '</strong>', styles['Normal']),Paragraph('<strong>' + str(c.carton.l3_9) + '</strong>', styles['Normal'])])
                else:
                    lineaCarton.append([Paragraph(str(c.carton.l3_1), styles['Normal']),Paragraph(str(c.carton.l3_2), styles['Normal']),Paragraph(str(c.carton.l3_3), styles['Normal']),Paragraph(str(c.carton.l3_4), styles['Normal']),Paragraph(str(c.carton.l3_5), styles['Normal']),Paragraph(str(c.carton.l3_6), styles['Normal']),Paragraph(str(c.carton.l3_7), styles['Normal']),Paragraph(str(c.carton.l3_8), styles['Normal']),Paragraph(str(c.carton.l3_9), styles['Normal'])])
                table = Table(lineaCarton)
                table.setStyle(TableStyle(
                    [
                        ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                        
                    ]
                ))
                cuerpo.append(table)
        #balotas 7mo premio
        header = Paragraph('SECUENCIA DE BALOTAS ', styles['Heading4'])
        cuerpo.append(header)
        CartonesJugando = []
        numeroTupla = ()
        bandera = 0
        padron = Jugada.objects.filter(partida_id = partida.id, premio = '7mo Premio').order_by('id')
        if padron:
            for c in padron:
                bandera += 1
                if bandera <= 10:
                    numeroTupla += (str(c.balota     ).zfill(2),)
                    if bandera == 10:
                        bandera = 0
                        CartonesJugando.append(numeroTupla)
                        numeroTupla = ()
            if numeroTupla:
                CartonesJugando.append(numeroTupla)

            table = Table(CartonesJugando)
            table.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (10, -1), 1, colors.dodgerblue),
                    
                ]
            ))
            cuerpo.append(table)

    doc.build(cuerpo)
    response.write(buff.getvalue())
    buff.close()
    return response

import os
from django.conf import settings
from reportlab.lib.units import mm
from sorl.thumbnail import get_thumbnail
def ImprimirCartonesBlank(request,id):
    print "Genero el PDF"
    inicioC = int(request.GET['inicio'])
    finC = int(request.GET['fin']) + 1
    print inicioC, finC
    partida = Partida.objects.get(pk=id)
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte-Partidas.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=A4,
                            rightMargin=20,
                            leftMargin=20,
                            topMargin=20,
                            bottomMargin=20,

                 )
    header = []
    allclientes = []
    styles = getSampleStyleSheet()
    
    #estlos cabecera
    stylesCabezera = ParagraphStyle(name='Heading2', alignment = TA_LEFT, spaceBefore=3, )
    stylesN = ParagraphStyle(name='Heading2', alignment = TA_RIGHT, spaceBefore=4, )
    stylesCell = ParagraphStyle(name='Normal', alignment = TA_LEFT, spaceBefore=1, )
    stylesBlanco = ParagraphStyle(name='Heading2', alignment = TA_LEFT, spaceBefore=7, )
    #imagen de la partida
    imagen_file = os.path.join(settings.STATIC_ROOT, str('img/fondoBlanco.jpg'))
    image = Image(imagen_file)
    header.append(image)
    # #organiza la partida
    # header.append(Paragraph("Organiza: " + str(partida), stylesCabezera))

    # #sorteo de la partida
    # header.append(Paragraph("Sorteo: " + partida.sorteo, stylesCabezera))

    # #transmite
    # header.append(Paragraph("Transmitido por: " + partida.transmite + "  Precio: " + partida.precio, stylesCabezera))

    carton = []
    for i in range(inicioC,finC):
        carton += header
        numCarton = Impresion.objects.get(partida_id = id,numeroCarton=i)
        Impresion.objects.filter(partida_id = id,numeroCarton=i).update(participa='S')
        CartonImprimir = Carton.objects.get(pk=numCarton.carton_id)
        carton.append(Paragraph("<strong><font size=14>CARTON: " + str(i).zfill(5) + "</font></strong>", stylesN))
        barcode128 = code128.Code128(str(numCarton.id),barHeight=.3*inch,barWidth = 1.2)
        #linea 1
        headings = ('                       ', CartonImprimir.l1_1,CartonImprimir.l1_2,CartonImprimir.l1_3,CartonImprimir.l1_4,CartonImprimir.l1_5,CartonImprimir.l1_6,CartonImprimir.l1_7,CartonImprimir.l1_8,CartonImprimir.l1_9)
        premios = []

        #Cudriculas de premios para la partida
        premios += [(Paragraph("<strong><font size=1>" + "</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]
        premios += [(Paragraph("<strong><font size=1>" + "</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]
        premios += [(Paragraph("<strong><font size=1>" + "</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]
        premios += [(Paragraph("<strong><font size=1>" + "</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]
        premios += [(Paragraph("<strong><font size=1>" + "</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]
        premios += [(Paragraph("<strong><font size=1>" + "</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]

        t = Table([headings] + premios,colWidths=(35*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm),rowHeights=(8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm),hAlign='LEFT')
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (9, -1), 1, HexColor('#FFFFFF')),
                ('BACKGROUND', (0, 0), (0, -1), HexColor('#FFFFFF')),
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                ('FONTSIZE', (0, 0), (-1, 0), 16),
                ('FONTSIZE', (0, 0), (0, -1), 14),
                ('TEXTFONT', (0, 0), (-1, 0), 'Helvetica-BoldOblique'),
                ('TEXTFONT', (0, 0), (0, -1), 'Helvetica-BoldOblique'), 
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                ('TOPPADDING', (0, 0), (-1, 0), 5), 
                ('BOTTOMPADDING', (0, 0), (0, -1), 10), 
                ('TOPPADDING', (0, 0), (0, -1), 5), 
            ]
        ))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(t)
        cantL = 0
        #linea 2
        if partida.lineasEnJuego >= 2:
            cantL = 2
            headings = ('                       ', CartonImprimir.l2_1,CartonImprimir.l2_2,CartonImprimir.l2_3,CartonImprimir.l2_4,CartonImprimir.l2_5,CartonImprimir.l2_6,CartonImprimir.l2_7,CartonImprimir.l2_8,CartonImprimir.l2_9)

            t = Table([headings] + premios,colWidths=(35*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm),rowHeights=(8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm),hAlign='LEFT')
            t.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (9, -1), 1, HexColor('#FFFFFF')),
                    ('BACKGROUND', (0, 0), (0, -1), HexColor('#FFFFFF')),
                    ('BACKGROUND', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                    ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('FONTSIZE', (0, 0), (-1, 0), 16),
                    ('FONTSIZE', (0, 0), (0, -1), 14),
                    ('TEXTFONT', (0, 0), (-1, 0), 'Helvetica-BoldOblique'),
                    ('TEXTFONT', (0, 0), (0, -1), 'Helvetica-BoldOblique'), 
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                    ('TOPPADDING', (0, 0), (-1, 0), 5), 
                    ('BOTTOMPADDING', (0, 0), (0, -1), 10), 
                    ('TOPPADDING', (0, 0), (0, -1), 5), 
                ]
            ))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(t)

        #linea 3
        if partida.lineasEnJuego >= 3:
            cantL = 3
            headings = ('                       ', CartonImprimir.l3_1,CartonImprimir.l3_2,CartonImprimir.l3_3,CartonImprimir.l3_4,CartonImprimir.l3_5,CartonImprimir.l3_6,CartonImprimir.l3_7,CartonImprimir.l3_8,CartonImprimir.l3_9)

            t = Table([headings] + premios,colWidths=(35*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm,12*mm),rowHeights=(8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm),hAlign='LEFT')

            t.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (9, -1), 1, HexColor('#FFFFFF')),
                    ('BACKGROUND', (0, 0), (0, -1), HexColor('#FFFFFF')),
                    ('BACKGROUND', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                    ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('FONTSIZE', (0, 0), (-1, 0), 16),
                    ('FONTSIZE', (0, 0), (0, -1), 14),
                    ('TEXTFONT', (0, 0), (-1, 0), 'Helvetica-BoldOblique'),
                    ('TEXTFONT', (0, 0), (0, -1), 'Helvetica-BoldOblique'), 
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                    ('TOPPADDING', (0, 0), (-1, 0), 5), 
                    ('BOTTOMPADDING', (0, 0), (0, -1), 10), 
                    ('TOPPADDING', (0, 0), (0, -1), 5), 
                ]
            ))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(t)
        """
        CUANDO LAS CONDICIONES ESTAN ALANTE PERO PASARON ATRAS
        #condiciones        
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph("Condiciones + Reglas de Juego: " + partida.condiciones, stylesCabezera))
        """
        carton.append(Paragraph(" ", stylesBlanco))
        if cantL == 2:
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(Paragraph(" ", stylesBlanco))

        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph(" ", stylesBlanco))

        #CODIGO DE BARRA
        #CODIGO DE BARRA
        carton.append(barcode128)
        carton.append(Paragraph("<strong>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;" + str(i).zfill(5) + "</strong>", stylesBlanco))

        #SALTO E PAGINA
        carton.append(PageBreak())

    doc.build(carton)
    response.write(buff.getvalue())
    buff.close()
    return response


#CARTON LIMPIO

def ImprimirCartones(request,id):
    print "Genero el PDF"
    inicioC = int(request.GET['inicio'])
    finC = int(request.GET['fin']) + 1
    print inicioC, finC
    partida = Partida.objects.get(pk=id)
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte-Partidas.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=A4,
                            rightMargin=20,
                            leftMargin=20,
                            topMargin=20,
                            bottomMargin=20,

                 )
    header = []
    allclientes = []
    styles = getSampleStyleSheet()
    
    #estlos cabecera
    stylesCabezera = ParagraphStyle(name='Heading2', alignment = TA_CENTER, spaceBefore=3, )
    stylesN = ParagraphStyle(name='Heading2', alignment = TA_RIGHT, spaceBefore=4, )
    stylesCell = ParagraphStyle(name='Normal', alignment = TA_CENTER, spaceBefore=1, )
    stylesBlanco = ParagraphStyle(name='Heading2', alignment = TA_CENTER, spaceBefore=7, )
    #imagen de la partida
    image = Image(partida.thumb)
    header.append(image)
    # #organiza la partida
    # header.append(Paragraph("Organiza: " + str(partida), stylesCabezera))

    # #sorteo de la partida
    # header.append(Paragraph("Sorteo: " + partida.sorteo, stylesCabezera))

    # #transmite
    # header.append(Paragraph("Transmitido por: " + partida.transmite + "  Precio: " + partida.precio, stylesCabezera))

    carton = []
    for i in range(inicioC,finC):
        carton += header
        numCarton = Impresion.objects.get(partida_id = id,numeroCarton=i)
        Impresion.objects.filter(partida_id = id,numeroCarton=i).update(participa='S')
        CartonImprimir = Carton.objects.get(pk=numCarton.carton_id)
        carton.append(Paragraph("<strong>CARTON: " + str(i).zfill(5) + "</strong>", stylesN))
        barcode128 = code128.Code128(str(numCarton.id),barHeight=.3*inch,barWidth = 1.2)
        #linea 1
        headings = ('        PREMIOS        ', CartonImprimir.l1_1,CartonImprimir.l1_2,CartonImprimir.l1_3,CartonImprimir.l1_4,CartonImprimir.l1_5,CartonImprimir.l1_6,CartonImprimir.l1_7,CartonImprimir.l1_8,CartonImprimir.l1_9)
        premios = []

        #Cudriculas de premios para la partida
        if partida.premio_1:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_1 +"</font></strong>", stylesCell),'            ','            ','            ','            ','            ','            ','            ','            ','            ')]
        if partida.premio_2:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_2 +"</font></strong>", stylesCell),'        ','        ','        ','        ','    ','    ','    ','    ','    ')]
        if partida.premio_3:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_3 +"</font></strong>", stylesCell),'    ','    ','    ','    ','    ','    ','    ','    ','    ')]
        if partida.premio_4:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_4 +"</font></strong>", stylesCell),'    ','    ','    ','    ','    ','    ','    ','    ','    ')]
        if partida.premio_5:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_5 +"</font></strong>", stylesCell),'    ','    ','    ','    ','    ','    ','    ','    ','    ')]
        if partida.premio_6:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_6 +"</font></strong>", stylesCell),'    ','    ','    ','    ','    ','    ','    ','    ','    ')]
        if partida.premio_7:
            premios += [(Paragraph("<strong><font size=13>" + partida.monto_7 +"</font></strong>", stylesCell),'    ','    ','    ','    ','    ','    ','    ','    ','    ')]

        t = Table([headings] + premios,rowHeights=(8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm))
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (9, -1), 1, HexColor(partida.t1l)),
                ('BACKGROUND', (0, 0), (0, -1), HexColor(partida.t1f)),
                ('BACKGROUND', (0, 0), (-1, 0), HexColor(partida.t1f)),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                ('FONTSIZE', (0, 0), (-1, 0), 16),
                ('FONTSIZE', (0, 0), (0, -1), 14),
                ('TEXTFONT', (0, 0), (-1, 0), 'Helvetica-BoldOblique'),
                ('TEXTFONT', (0, 0), (0, -1), 'Helvetica-BoldOblique'), 
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                ('TOPPADDING', (0, 0), (-1, 0), 5), 
                ('BOTTOMPADDING', (0, 0), (0, -1), 10), 
                ('TOPPADDING', (0, 0), (0, -1), 5), 
            ]
        ))
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(t)

        #linea 2
        if partida.lineasEnJuego >= 2:
            headings = ('        PREMIOS        ', CartonImprimir.l2_1,CartonImprimir.l2_2,CartonImprimir.l2_3,CartonImprimir.l2_4,CartonImprimir.l2_5,CartonImprimir.l2_6,CartonImprimir.l2_7,CartonImprimir.l2_8,CartonImprimir.l2_9)

            t = Table([headings] + premios,rowHeights=(8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm))
            t.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (9, -1), 1, HexColor(partida.t2l)),
                    ('BACKGROUND', (0, 0), (0, -1), HexColor(partida.t2f)),
                    ('BACKGROUND', (0, 0), (-1, 0), HexColor(partida.t2f)),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('FONTSIZE', (0, 0), (-1, 0), 16),
                    ('FONTSIZE', (0, 0), (0, -1), 14),
                    ('TEXTFONT', (0, 0), (-1, 0), 'Helvetica-BoldOblique'),
                    ('TEXTFONT', (0, 0), (0, -1), 'Helvetica-BoldOblique'), 
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                    ('TOPPADDING', (0, 0), (-1, 0), 5), 
                    ('BOTTOMPADDING', (0, 0), (0, -1), 10), 
                    ('TOPPADDING', (0, 0), (0, -1), 5), 
                ]
            ))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(t)

        #linea 3
        if partida.lineasEnJuego >= 3:
            headings = ('        PREMIOS        ', CartonImprimir.l3_1,CartonImprimir.l3_2,CartonImprimir.l3_3,CartonImprimir.l3_4,CartonImprimir.l3_5,CartonImprimir.l3_6,CartonImprimir.l3_7,CartonImprimir.l3_8,CartonImprimir.l3_9)

            t = Table([headings] + premios,rowHeights=(8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm,8*mm))
            t.setStyle(TableStyle(
                [
                    ('GRID', (0, 0), (9, -1), 1, HexColor(partida.t3l)),
                    ('BACKGROUND', (0, 0), (0, -1), HexColor(partida.t3f)),
                    ('BACKGROUND', (0, 0), (-1, 0), HexColor(partida.t3f)),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('FONTSIZE', (0, 0), (-1, 0), 16),
                    ('FONTSIZE', (0, 0), (0, -1), 14),
                    ('TEXTFONT', (0, 0), (-1, 0), 'Helvetica-BoldOblique'),
                    ('TEXTFONT', (0, 0), (0, -1), 'Helvetica-BoldOblique'), 
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10), 
                    ('TOPPADDING', (0, 0), (-1, 0), 5), 
                    ('BOTTOMPADDING', (0, 0), (0, -1), 10), 
                    ('TOPPADDING', (0, 0), (0, -1), 5), 
                ]
            ))
            carton.append(Paragraph(" ", stylesBlanco))
            carton.append(t)
        """
        CUANDO LAS CONDICIONES ESTAN ALANTE PERO PASARON ATRAS
        #condiciones        
        carton.append(Paragraph(" ", stylesBlanco))
        carton.append(Paragraph("Condiciones + Reglas de Juego: " + partida.condiciones, stylesCabezera))
        """
        carton.append(Paragraph(" ", stylesBlanco))

        #CODIGO DE BARRA
        carton.append(barcode128)

        #SALTO E PAGINA
        carton.append(PageBreak())

    doc.build(carton)
    response.write(buff.getvalue())
    buff.close()
    return response



#Nuestra clase hereda de la vista genérica TemplateView
class ReportePartidasExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = Cliente.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        # ws['B1'] = 'REPORTE DE PERSONAS'
        # #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        # ws.merge_cells('B1:E1')
        # #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A1'] = 'P. ID'
        ws['B1'] = 'P. Descripcion'
        ws['C1'] = 'S. ID'
        ws['D1'] = 'S. Nombre'       
        ws['E1'] = 'Asignacion'       
        cont=2
        
        dt = datetime.datetime.strptime(request.GET['v'] + " 00:00", "%m/%d/%Y %H:%M")
        #Recorremos el conjunto de las Partidas por la fecha parametro del get
        partidas = Partida.objects.filter(fecha = dt);
        for partida in partidas:
            for sala in Sala.objects.filter(activa='Si'):
                ws.cell(row=cont,column=1).value = partida.id
                ws.cell(row=cont,column=2).value = partida.descripcion
                ws.cell(row=cont,column=3).value = sala.id
                ws.cell(row=cont,column=4).value = sala.nombre  
                ws.cell(row=cont,column=5).value = 0
                cont = cont + 1
        # #Establecemos el nombre del archivo
        nombre_archivo = "Asignar.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response




#Nuestra clase hereda de la vista genérica TemplateView
class MoeloPartida(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = Cliente.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        # ws['B1'] = 'REPORTE DE PERSONAS'
        # #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        # ws.merge_cells('B1:E1')
        # #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A1'] = 'Fecha'
        ws['B1'] = 'Descripcion'
        ws['C1'] = 'L1'
        ws['D1'] = 'L1-L2'
        ws['E1'] = 'L1-L3'       
        ws['F1'] = 'L2'       
        ws['G1'] = 'L2-L3'       
        ws['H1'] = 'L3'       
        ws['I1'] = 'BINGO'       
        ws['J1'] = 'LA'       
        ws['K1'] = 'REVANCHA'       
        ws['L1'] = 'REBINGO'       
        # #Establecemos el nombre del archivo
        nombre_archivo = "FormatoPartidas.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
